import streamlit as st
import pandas as pd
import os
import re
import zipfile
import subprocess
from docxtpl import DocxTemplate
from openpyxl import load_workbook
from num2words import num2words

st.title("Salary Payslip Generator")

excel_file = st.file_uploader("Upload Excel Salary Sheet", type=["xlsx"])
template_file = st.file_uploader("Upload Word Payslip Template", type=["docx"])

if excel_file and template_file:

    os.makedirs("payslips", exist_ok=True)

    with open("salary.xlsx", "wb") as f:
        f.write(excel_file.read())

    with open("template.docx", "wb") as f:
        f.write(template_file.read())

    wb = load_workbook("salary.xlsx")
    ws = wb.active

    title = str(ws["A1"].value)

    match = re.search(r"month of (.*)", title, re.IGNORECASE)

    if match:
        month = match.group(1)
        st.success(f"Detected Month: {month}")

    else:
        st.warning("Month not detected in Excel")

        months = [
            "January","February","March","April","May","June",
            "July","August","September","October","November","December"
        ]

        selected_month = st.selectbox("Select Month", months)
        year = st.text_input("Enter Year")

        month = f"{selected_month} {year}"

    temp_df = pd.read_excel("salary.xlsx", header=None)

    header_row = None

    for i in range(len(temp_df)):
        row = temp_df.iloc[i].astype(str).str.upper().tolist()

        if "NAME OF EMPLOYEE" in row or "EMPLOYEE NAME" in row:
            header_row = i
            break

    if header_row is None:
        st.error("Header row not detected")
        st.stop()

    df = pd.read_excel("salary.xlsx", header=header_row)

    df.columns = df.columns.astype(str).str.strip().str.upper()

    st.write("Detected Columns:", df.columns.tolist())

    pdf_files = []

    def get_value(row, cols):

        for col in cols:

            if col in row.index:

                val = row[col]

                if pd.isna(val):
                    return 0

                return float(val)

        return 0

    if st.button("Generate Payslips"):

        for _, row in df.iterrows():

            employee = str(row.get("NAME OF EMPLOYEE", "Employee")).strip()

            basic = get_value(row, ["BASIC", "BASIC SALARY"])

            if basic == 0:
                st.warning(f"Skipping employee missing BASIC: {employee}")
                continue

            hra = get_value(row, ["HRA"])
            spl = get_value(row, ["SPL. ALL.", "SPECIAL ALLOWANCE"])
            lta = get_value(row, ["LTA"])
            variable = get_value(row, ["VARIABLE PAY"])

            pf = get_value(row, ["PF"])
            pt = get_value(row, ["PT"])
            employer = get_value(row, ["EMPLOYER CONTRIBUTION", "EMPLOYER PF"])

            gross = get_value(row, ["GROSS"])
            gross_dedns = get_value(row, ["GROSS DEDNS.", "GROSS DEDUCTIONS"])
            total_deduction = get_value(row, ["TOTAL DEDUCTIONS"])
            net_pay = get_value(row, ["NET PAY"])

            if gross == 0:
                gross = basic + hra + spl + lta + variable

            if gross_dedns == 0:
                gross_dedns = pf + pt

            if total_deduction == 0:
                total_deduction = gross_dedns + employer

            if net_pay == 0:
                net_pay = gross - gross_dedns - employer

            net_words = num2words(int(net_pay), lang="en_IN").title()

            context = {
                "Month": month,
                "Employee_Name": employee,
                "EmpNo": row.get("EMP.NO.", ""),
                "Basic": basic,
                "HRA": hra,
                "Spl_All": spl,
                "LTA": lta,
                "PF": pf,
                "PT": pt,
                "Employer_Contribution": employer,
                "GROSS": gross,
                "Gross_Dedns": gross_dedns,
                "Total_Deduction": total_deduction,
                "NetPay": net_pay,
                "NetPayWords": net_words
            }

            doc = DocxTemplate("template.docx")
            doc.render(context)

            emp = employee.replace(" ", "_")

            temp_docx = f"{emp}.docx"

            doc.save(temp_docx)

            subprocess.run([
                "libreoffice",
                "--headless",
                "--convert-to",
                "pdf",
                temp_docx,
                "--outdir",
                "payslips"
            ])

            generated_pdf = f"payslips/{emp}.pdf"

            final_pdf = f"payslips/payslip_{emp}_{month.replace(' ','_')}.pdf"

            if os.path.exists(generated_pdf):
                os.rename(generated_pdf, final_pdf)
                pdf_files.append(final_pdf)

            os.remove(temp_docx)

        st.success("Payslips Generated Successfully")

        zip_name = "payslips_pdf_only.zip"

        with zipfile.ZipFile(zip_name, 'w') as z:

            for pdf in pdf_files:
                z.write(pdf, os.path.basename(pdf))

        with open(zip_name, "rb") as f:
            st.download_button(
                "Download ZIP",
                f,
                file_name="payslips_pdf_only.zip"
            )
